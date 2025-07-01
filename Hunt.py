from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import time

class Hunt():
    def process_data(self, input_path, output_path):
        try:
            # Load Excel with Property IDs
            input_wb = load_workbook(input_path)
            input_ws = input_wb.active

            # Create output Excel
            output_wb = Workbook()
            output_ws = output_wb.active
            output_ws.append(["prop_id", "owner_name", "owner_address"])

            # Set up Selenium Chrome with options
            options = webdriver.ChromeOptions()
            options.add_argument("disable-infobars")
            options.add_argument("--disable-extensions")
            options.add_argument("--start-maximized")
            driver = webdriver.Chrome(options=options)

            # Load the URL to the search page
            driver.get("https://esearch.hunt-cad.org/")
            print("Opened the website successfully.")

            # Loop through each Property ID in Excel
            for row in input_ws.iter_rows(min_row=2, values_only=True):
                prop_id = str(row[0]).strip()
                if not prop_id or prop_id == "None":
                    continue                            # Empty or invalid property ID. Skipping. So continue.

                print(f"Searching for ID: {prop_id}")

                # Click "By ID" tab
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="home-page-tabs"]/li[3]/a'))
                ).click()

                # Enter ID into search bar
                search_input = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.ID, "PropertyId"))
                )
                search_input.clear()
                search_input.send_keys(prop_id)
                search_input.send_keys(Keys.RETURN)

                # Click the first result row
                try:
                    WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="resultListDiv"]/tr[2]'))
                    ).click()
                except Exception as e:
                    print(f"❌ No clickable result for {prop_id}: {e}")
                    output_ws.append([prop_id, "", ""])
                    # Return to the search page
                    driver.get("https://esearch.hunt-cad.org/")  
                    continue

                # Wait for detail page to load
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "detail-page"))
                )

                # Scrape owner info
                try:
                    # Locate and extract the owner name
                    owner_name = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="detail-page"]/div[2]/div[1]/div/table/tbody/tr[13]/td'))
                    ).text
                    # Locate and extract the owner address
                    owner_address = driver.find_element(
                        By.XPATH, '//*[@id="detail-page"]/div[2]/div[1]/div/table/tbody/tr[15]/td'
                    ).text
                except Exception as e:
                    owner_name = ""
                    owner_address = ""
                    print(f"Error scraping data for {prop_id}: {e}")

                # Save results to Excel
                output_ws.append([prop_id, owner_name, owner_address])

                # Return to the search page
                driver.get("https://esearch.hunt-cad.org/")  

            # Output the extracted information into Excel
            output_wb.save(output_path)
            print("Results saved successfully.")
            return "Success"

        except Exception as e:
            print(f"An error occurred: {e}")
            return str(e)

        finally:
            try:
                driver.quit()
            except:
                pass
