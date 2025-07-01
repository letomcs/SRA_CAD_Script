from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import time

class Panola():
    def process_data(self, input_path, output_path):
        try:
            # Load Excel with Property IDs
            input_wb = load_workbook(input_path)
            input_ws = input_wb.active

            # Create output Excel
            output_wb = Workbook()
            output_ws = output_wb.active
            output_ws.append(["prop_id", "owner_name", "owner_address"])

            # Set up Selenium Chrome with options
            options = webdriver.ChromeOptions()
            options.add_argument("disable-infobars")
            options.add_argument("--disable-extensions")
            options.add_argument("--start-maximized")
            driver = webdriver.Chrome(options=options)

            driver.get("https://www.panolacad.org/home/Custom?custurl=/home")
            print("Opened the website successfully.")

            for row in input_ws.iter_rows(min_row=2, values_only=True):
                prop_id = str(row[0]).strip()
                if not prop_id or prop_id == "None":
                    continue                            # Empty or invalid property ID. Skipping. So continue.

                print(f"Searching for ID: {prop_id}")

                # Enter ID into search bar
                search_input = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'Keyword'))
                )
                search_input.clear()
                search_input.send_keys(prop_id)
                search_input.send_keys(Keys.RETURN)

                # # Wait for result table to load
                # WebDriverWait(driver, 10).until(
                #     EC.presence_of_element_located((By.XPATH, '//*[@id="resultListDiv"]'))
                # )

                # Click the first result row
                try:
                    WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="id_search_list"]/tbody/tr[1]'))
                    ).click()
                except Exception as e:
                    print(f"❌ No clickable result for {prop_id}: {e}")
                    output_ws.append([prop_id, "", ""])
                    # Return to the search page
                    driver.get("https://www.panolacad.org/home/Custom?custurl=/home")  
                    continue

                # Scrape owner info
                try:
                    # Locate and extract the owner name
                    owner_name = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="owner-card"]/div[2]/table/tbody/tr[2]/td[2]'))
                    ).text
                    # Locate and extract the owner address
                    owner_address = driver.find_element(
                        By.XPATH, '//*[@id="owner-card"]/div[2]/table/tbody/tr[4]/td[2]'
                    ).text
                except Exception as e:
                    owner_name = ""
                    owner_address = ""
                    print(f"Error scraping data for {prop_id}: {e}")

                # Save results to Excel
                output_ws.append([prop_id, owner_name, owner_address])

                # Return to the search page
                driver.get("https://www.panolacad.org/home/Custom?custurl=/home")  

            output_wb.save(output_path)
            print("Results saved successfully.")
            return "Success"

        except Exception as e:
            print(f"An error occurred: {e}")
            return str(e)

        finally:
            try:
                driver.quit()
            except:
                pass
